from __future__ import annotations

import os
import re
import ssl
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.parse import parse_qsl, unquote, urlparse
from uuid import uuid4

import pymysql
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT_PATH = PROJECT_ROOT / "dataset.xlsx"
DEFAULT_DATABASE = "review_intelligence"
DEFAULT_IMPORT_BATCH_SIZE = 100
DEFAULT_IMPORT_LIMIT = 20000
MAX_IMPORT_BATCH_SIZE = 100
REPLACE_EXISTING_IMPORT = os.environ.get("IMPORT_REPLACE_EXISTING", "true") != "false"
HTML_TAG_RE = re.compile(r"<[^>]*>")
CONTROL_CHARACTER_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
VARCHAR_LIMITS = {
    "review_id": 128,
    "product_id": 128,
    "customer_id": 128,
    "customer_name": 255,
}
REQUIRED_COLUMNS = [
    "review_id",
    "product_id",
    "customer_id",
    "review_title",
    "review_written_date",
    "customer_name",
    "review_from_title",
    "review_text",
    "helpful_count",
    "out_of_helpful_count",
    "customer_review_rating",
    "number_of_comments",
    "amazon_verified_purchase",
    "amazon_vine_program_review",
]


def load_env_file(path: Path) -> None:
    if not path.exists():
        return

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()

        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()

        if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
            value = value[1:-1]

        os.environ.setdefault(key, value)


def load_local_env() -> None:
    node_env = os.environ.get("NODE_ENV", "development")

    for name in (
        f".env.{node_env}.local",
        ".env.local",
        f".env.{node_env}",
        ".env",
    ):
        load_env_file(PROJECT_ROOT / name)


def clean_string(value: Any) -> str | None:
    if value is None:
        return None

    if isinstance(value, (date, datetime)):
        return value.isoformat()

    text = unicodedata.normalize("NFKC", str(value)).strip()

    if not text or text.lower() in {"null", "nan", "none"}:
        return None

    if len(text) >= 2 and text.startswith("'") and text.endswith("'"):
        text = text[1:-1].replace("\\'", "'")

    cleaned = (
        text.replace("\\r\\n", "\n")
        .replace("\\n", "\n")
        .replace("\\r", "\n")
        .replace("\r\n", "\n")
        .replace("\r", "\n")
        .strip()
    )

    cleaned = HTML_TAG_RE.sub(" ", cleaned)
    cleaned = CONTROL_CHARACTER_RE.sub(" ", cleaned)
    cleaned = re.sub(r"[ \t\f\v]+", " ", cleaned)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned).strip()

    return cleaned or None


def bounded_string(value: Any, max_length: int) -> str | None:
    text = clean_string(value)

    if text is None:
        return None

    return text[:max_length]


def normalized_review_key(title: Any, text: Any) -> str | None:
    cleaned_text = clean_string(text)

    if not cleaned_text:
        return None

    cleaned_title = clean_string(title) or ""
    normalized = unicodedata.normalize("NFKC", f"{cleaned_title} {cleaned_text}").lower()
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()

    return normalized or None


def to_unsigned_integer(value: Any) -> int | None:
    text = clean_string(value)

    if text is None:
        return None

    try:
        number = int(float(text))
    except ValueError:
        return None

    if number < 0:
        return None

    return number


def to_rating(value: Any) -> int | None:
    rating = to_unsigned_integer(value)

    if rating is None or rating < 1 or rating > 5:
        return None

    return rating


def to_boolean(value: Any) -> int | None:
    text = clean_string(value)

    if text is None:
        return None

    normalized = text.lower()

    if normalized in {"1", "true", "yes", "y"}:
        return 1

    if normalized in {"0", "false", "no", "n"}:
        return 0

    return None


def to_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    text = clean_string(value)

    if text is None:
        return None

    if len(text) >= 10 and text[4] == "-" and text[7] == "-":
        return text[:10]

    for fmt in ("%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue

    return None


def normalize_certificate(value: str) -> str:
    return value.strip().replace("\\n", "\n").replace("\\", "\n")


def create_ssl_context() -> ssl.SSLContext:
    ca = os.environ.get("MYSQL_CA_CERT", "").strip()
    ca_path = os.environ.get("MYSQL_CA_CERT_PATH", "").strip()
    reject_unauthorized = os.environ.get("MYSQL_SSL_REJECT_UNAUTHORIZED") != "false"

    if ca:
        return ssl.create_default_context(cadata=normalize_certificate(ca))

    if ca_path:
        return ssl.create_default_context(cafile=ca_path)

    if reject_unauthorized:
        return ssl.create_default_context()

    return ssl._create_unverified_context()


def parse_database_url(database_url: str) -> dict[str, Any]:
    parsed = urlparse(database_url)

    if parsed.scheme not in {"mysql", "mysql2"}:
        raise ValueError("DATABASE_URL must use the mysql:// scheme.")

    query = dict(parse_qsl(parsed.query))

    return {
        "host": parsed.hostname,
        "port": parsed.port or 3306,
        "user": unquote(parsed.username or ""),
        "password": unquote(parsed.password or ""),
        "database": os.environ.get("MYSQL_DATABASE") or parsed.path.lstrip("/")
        or DEFAULT_DATABASE,
        "charset": "utf8mb4",
        "cursorclass": pymysql.cursors.Cursor,
        "autocommit": True,
        "connect_timeout": 10,
        "read_timeout": 300,
        "write_timeout": 300,
        "ssl": create_ssl_context() if query.get("ssl-mode") != "DISABLED" else None,
    }


def validate_columns(columns: list[str]) -> None:
    missing = [column for column in REQUIRED_COLUMNS if column not in columns]

    if missing:
        raise ValueError(f"Missing required column(s): {', '.join(missing)}")


def map_review_row(
    row: dict[str, Any],
    dataset_id: str,
    source_row: int,
) -> tuple[Any, ...] | None:
    review_text = clean_string(row.get("review_text"))

    if not review_text:
        return None

    return (
        dataset_id,
        source_row,
        bounded_string(row.get("review_id"), VARCHAR_LIMITS["review_id"]),
        bounded_string(row.get("product_id"), VARCHAR_LIMITS["product_id"]),
        bounded_string(row.get("customer_id"), VARCHAR_LIMITS["customer_id"]),
        clean_string(row.get("review_title")),
        review_text,
        to_date(row.get("review_written_date")),
        bounded_string(row.get("customer_name"), VARCHAR_LIMITS["customer_name"]),
        clean_string(row.get("review_from_title")),
        to_unsigned_integer(row.get("helpful_count")),
        to_unsigned_integer(row.get("out_of_helpful_count")),
        to_rating(row.get("customer_review_rating")),
        to_unsigned_integer(row.get("number_of_comments")),
        to_boolean(row.get("amazon_verified_purchase")),
        to_boolean(row.get("amazon_vine_program_review")),
        None,
    )


def insert_batch(cursor: pymysql.cursors.Cursor, rows: list[tuple[Any, ...]]) -> None:
    if not rows:
        return

    cursor.executemany(
        """
        INSERT INTO reviews
          (
            dataset_id,
            source_row,
            review_id,
            product_id,
            customer_id,
            review_title,
            review_text,
            review_written_date,
            customer_name,
            review_from_title,
            helpful_count,
            out_of_helpful_count,
            customer_review_rating,
            number_of_comments,
            amazon_verified_purchase,
            amazon_vine_program_review,
            raw_metadata
          )
        VALUES
          (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """,
        rows,
    )


def delete_replaced_datasets(
    cursor: pymysql.cursors.Cursor,
    active_dataset_id: str,
) -> None:
    if not REPLACE_EXISTING_IMPORT:
        return

    cursor.execute(
        """
        DELETE FROM datasets
        WHERE id <> %s
        """,
        (active_dataset_id,),
    )


def mark_dataset_failed(database_url: str, dataset_id: str, error: Exception) -> None:
    try:
        connection = pymysql.connect(**parse_database_url(database_url))

        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    UPDATE datasets
                    SET status = 'failed', error_message = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                    """,
                    (str(error), dataset_id),
                )
        finally:
            connection.close()
    except Exception as marker_error:
        print(
            f"Could not mark dataset {dataset_id} as failed: {marker_error}",
            file=sys.stderr,
        )


def main() -> int:
    load_local_env()

    database_url = os.environ.get("DATABASE_URL")

    if not database_url:
        print("DATABASE_URL is required.", file=sys.stderr)
        return 1

    input_path = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else DEFAULT_INPUT_PATH
    configured_batch_size = int(
        os.environ.get("IMPORT_BATCH_SIZE") or DEFAULT_IMPORT_BATCH_SIZE
    )
    batch_size = min(configured_batch_size, MAX_IMPORT_BATCH_SIZE)
    import_limit = int(os.environ.get("IMPORT_LIMIT") or DEFAULT_IMPORT_LIMIT)
    dataset_id = str(uuid4())
    inserted_count = 0
    skipped_count = 0
    batch: list[tuple[Any, ...]] = []
    seen_review_keys: set[str] = set()
    dataset_created = False

    print(f"Reading {input_path}")

    if configured_batch_size != batch_size:
        print(
            f"Using import batch size {batch_size}; "
            f"configured value {configured_batch_size} is too large for safe inserts."
        )

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    columns = [str(column).strip() for column in next(rows)]

    validate_columns(columns)

    connection = pymysql.connect(**parse_database_url(database_url))

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO datasets
                  (id, source_filename, text_column, rating_column, row_count, status)
                VALUES
                  (%s, %s, %s, %s, 0, 'uploaded')
                """,
                (dataset_id, input_path.name, "review_text", "customer_review_rating"),
            )
            dataset_created = True

            for source_row, values in enumerate(rows, start=1):
                row = dict(zip(columns, values))
                review_key = normalized_review_key(
                    row.get("review_title"),
                    row.get("review_text"),
                )

                if review_key is None or review_key in seen_review_keys:
                    skipped_count += 1
                    continue

                seen_review_keys.add(review_key)
                mapped_row = map_review_row(row, dataset_id, source_row)

                if mapped_row is None:
                    skipped_count += 1
                    continue

                batch.append(mapped_row)

                if len(batch) >= batch_size:
                    insert_batch(cursor, batch)
                    inserted_count += len(batch)
                    batch.clear()
                    cursor.execute(
                        """
                        UPDATE datasets
                        SET row_count = %s, updated_at = CURRENT_TIMESTAMP
                        WHERE id = %s
                        """,
                        (inserted_count, dataset_id),
                    )
                    print(f"Imported {inserted_count} reviews...")

                if import_limit > 0 and inserted_count + len(batch) >= import_limit:
                    break

            insert_batch(cursor, batch)
            inserted_count += len(batch)

            cursor.execute(
                """
                UPDATE datasets
                SET row_count = %s, status = 'indexed', updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
                """,
                (inserted_count, dataset_id),
            )
            delete_replaced_datasets(cursor, dataset_id)

        print(f"Dataset {dataset_id} imported from {worksheet.title}.")
        print(f"Inserted {inserted_count} reviews. Skipped {skipped_count}.")
        return 0
    except Exception as exc:
        if dataset_created:
            mark_dataset_failed(database_url, dataset_id, exc)
        raise
    finally:
        workbook.close()
        connection.close()


if __name__ == "__main__":
    raise SystemExit(main())
